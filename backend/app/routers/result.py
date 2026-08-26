import io
import json
from datetime import datetime
from enum import Enum
from typing import List, Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from fastapi import APIRouter, Depends, HTTPException, Query, status, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from database.connection import get_db
from database.models import (
    ExamSession, StudentResult, Student, ExamShared, ExamPermission, User
)
from rate_limiting import limiter
from routers.auth import get_current_user
from services.results_service import (
    get_result_list,
    get_result_detail,
    manual_edit_result
)

class ResultStatus(str, Enum):
    VALID = "VALID"
    NEED_REVIEW = "NEED_REVIEW"
    MISSING_SBD = "MISSING_SBD"
    MISSING_TEST_CODE = "MISSING_TEST_CODE"
    UNKNOWN = "UNKNOWN"

class ResultSummary(BaseModel):
    total_submitted: int
    valid_count: int
    warning_count: int

class ResultItem(BaseModel):
    result_id: Optional[int]
    image_url: str
    student_code: Optional[str]
    student_name: Optional[str]
    test_code: Optional[str]
    total_score: float
    status: ResultStatus
    is_manual_override: bool
    warnings: List[str]
    batch_id: Optional[int]

class ResultListResponse(BaseModel):
    summary: ResultSummary
    items: List[ResultItem]

class QuestionDetail(BaseModel):
    question_number: int
    student_answer: Optional[str]
    correct_answer: Optional[str]
    is_correct: bool
    earned_score: float
    max_score: float

class ResultDetailResponse(BaseModel):
    result_id: int
    image_url: str
    student_code: Optional[str]
    student_name: Optional[str]
    test_code: Optional[str]
    total_score: float
    status: ResultStatus
    is_manual_override: bool
    questions: List[QuestionDetail]

class ManualEditRequest(BaseModel):
    student_code: Optional[str] = None
    test_code: Optional[str] = None
    answers: Optional[Dict[str,str]] = None

router = APIRouter(prefix="/api/v1", tags=["Results & Manual Edit"])

def check_session_permission(
    session_id: int,
    user_id: int,
    db: Session,
    required_permission: Optional[ExamPermission] = None
):

    session = db.query(ExamSession).filter(ExamSession.id == session_id).first()
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Đợt thi không tồn tại"
        )

    if session.created_by == user_id or session.exam.created_by == user_id:
        return session

    share_perm = db.query(ExamShared).filter(
        ExamShared.exam_id == session.exam_id,
        ExamShared.user_id == user_id
    ).first()

    if not share_perm:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Bạn không có quyền truy cập vào đợt thi này"
        )

    if required_permission:
        if required_permission == ExamPermission.EDITOR and share_perm.permission != ExamPermission.EDITOR:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Bạn cần quyền EDITOR để thực hiện thao tác này"
            )

    return session

@router.get("/sessions/{session_id}/results", summary="Lấy danh sách kết quả bài thi")
@limiter.limit("100/minute")
def get_results(
        request: Request,
        session_id: int,
        status_filter: str = Query("ALL", enum=["ALL", "WARNING_ONLY", "VALID_ONLY"]),
        search: Optional[str] = Query(None, description="Tìm theo SBD hoặc tên"),
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=100),
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    check_session_permission(session_id, current_user.id, db)

    try:
        data = get_result_list(
            session_id=session_id,
            db = db,
            status_filter = status_filter,
            search = search,
            page = page,
            limit = limit
        )
        return {"status": "success", "data": data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{result_id}", summary="Chi tiết 1 bài thi")
@limiter.limit("10/minute")
def get_result_detail_app(
        request: Request,
        result_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    result = db.query(StudentResult).filter(StudentResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy bài thi")

    check_session_permission(result.session_id, current_user.id, db)

    try:
        data = get_result_detail(result_id, db)
        return {"status": "success", "data": data}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/result/{result_id}", summary="Giáo viên chỉnh sửa tay")
@limiter.limit("10/minute")
def manual_edit(
        request: Request,
        result_id: int,
        edit_data: ManualEditRequest,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    result = db.query(StudentResult).filter(StudentResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Không tìm thấy bài thi")

    check_session_permission(result.session_id, current_user.id, db, required_permission=ExamPermission.EDITOR)

    try:
        data = manual_edit_result(
            result_id=result_id,
            edit_data=edit_data.dict(exclude_none=True),
            db=db,
            verified_by=current_user.id
        )
        return {"status": "success", "data": data}
    except ValueError as e:
        raise HTTPException(status_code=403, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/results/{result_id}", summary="Xóa kết quả bài thi này")
@limiter.limit("5/minute")
def delete_single_student_result(
        request: Request,
        result_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    result = db.query(StudentResult).filter(StudentResult.id == result_id).first()
    if not result:
        raise HTTPException(404, "Không tìm thấy bản ghi kết quả bài thi")

    check_session_permission(result.session_id, current_user.id, db, required_permission=ExamPermission.EDITOR)

    db.delete(result)
    db.commit()

    return {
        "status": "success",
        "message": f"Đã xóa thành công kết quả bài thi ID {result_id}"
    }

@router.delete("/sessions/{session_id}/results/clear", summary="Xóa toàn bộ kết quả bài chấm trong Đợt thi")
@limiter.limit("5/minute")
def clear_all_session_results(
        request: Request,
        session_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    session_obj = check_session_permission(session_id, current_user.id, db, required_permission=ExamPermission.EDITOR)

    deleted_count = db.query(StudentResult).filter(StudentResult.session_id == session_id).delete(synchronize_session=False)
    db.commit()

    return {
        "status": "success",
        "message": f"Đã xóa toàn bộ {deleted_count} kết quả bài chấm thuộc đợt thi '{session_obj.session_name}'",
        "deleted_count": deleted_count
    }

@router.get("/sessions/{session_id}/export-excel", summary="Xuất file Excel Bảng điểm tổng hợp")
@limiter.limit("100/minute")
def export_session_results_excel(
        request: Request,
        session_id: int,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    session_obj = check_session_permission(session_id, current_user.id, db)

    results = db.query(StudentResult).filter(StudentResult.session_id == session_id).all()
    if not results:
        raise HTTPException(
            404,
            "Đợt thi chưa có dữ liệu kết quả chấm điểm để xuất Excel"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng điểm tổng hợp"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    max_q = session_obj.total_questions or 40
    headers = ["STT", "Mã Học Sinh / SBD", "Họ và Tên", "Mã Đề", "Số Câu Đúng", "Tổng Số Câu", "Điểm Số", "Trạng Thái"]
    for q_idx in range(1, max_q + 1):
        headers.append(f"C{q_idx}")

    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
    # Màu cho đúng/sai
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    correct_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    wrong_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # ===== HÀM CHUYỂN ĐỔI TRẠNG THÁI =====
    def get_status_text(status, is_manual_override):
        if is_manual_override:
            return "Chấm tay"
        status_map = {
            "VALID": "Đã chấm",
            "NEED_REVIEW": "Cần xem lại",
            "PROCESSING": "Đang xử lý",
            "PENDING": "Đang xử lý",
            "GRADED": "Đã chấm",
            "COMPLETED": "Đã chấm",
        }
        return status_map.get(status, status or "Đang xử lý")

    for idx, res in enumerate(results, start=1):
        student_name = res.student.full_name if res.student else "Chưa định danh"
        student_code = res.student_code or (res.student.student_code if res.student else "N/A")
        final_score = res.manual_score if res.is_manual_override and res.manual_score is not None else res.score

        # ===== LẤY TRẠNG THÁI TIẾNG VIỆT =====
        status_text = get_status_text(res.status, res.is_manual_override)

        row_data = [
            idx,
            student_code,
            student_name,
            res.detected_test_code or "N/A",
            res.correct_count or 0,
            res.total_questions or max_q,
            round(final_score, 2) if final_score is not None else 0,
            status_text
        ]

        user_answers = res.manual_answers if res.is_manual_override and res.manual_answers else res.answers

        if isinstance(user_answers, str):
            try:
                user_answers = json.loads(user_answers)
            except:
                user_answers = {}

        if not isinstance(user_answers, dict):
            user_answers = {}

        for q_idx in range(1, max_q + 1):
            key = str(q_idx)
            question_data = user_answers.get(key, {})

            if isinstance(question_data, dict):
                choice = question_data.get('choice', '')
                is_correct = question_data.get('is_correct', False)
            else:
                choice = str(question_data) if question_data else ''
                is_correct = False

            if choice:
                row_data.append(choice)
            else:
                row_data.append('')

        ws.append(row_data)

        current_row = idx + 1

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border_thin
            if col_num in [1, 2, 4, 5, 6, 7, 8] or col_num > 8:
                cell.alignment = center_align

            if col_num > 8:
                value = cell.value
                if value:
                    q_idx = col_num - 8
                    key = str(q_idx)
                    question_data = user_answers.get(key, {})

                    if isinstance(question_data, dict):
                        is_correct = question_data.get('is_correct', False)
                    else:
                        is_correct = False

                    if is_correct:
                        cell.fill = correct_fill
                        cell.font = Font(color="000000", bold=True)
                    else:
                        cell.fill = wrong_fill
                        cell.font = Font(color="FFFFFF", bold=True)

    # ===== ĐỊNH DẠNG CỘT ĐIỂM =====
    for row in range(2, len(results) + 2):
        score_cell = ws.cell(row=row, column=7)
        if score_cell.value is not None:
            try:
                score_val = float(score_cell.value)
                if score_val >= 8:
                    score_cell.font = Font(color="008000", bold=True)
                elif score_val >= 5:
                    score_cell.font = Font(color="0000FF")
                else:
                    score_cell.font = Font(color="FF0000")
            except:
                pass

    # ===== ĐỊNH DẠNG CỘT TRẠNG THÁI =====
    for row in range(2, len(results) + 2):
        status_cell = ws.cell(row=row, column=8)
        status_value = status_cell.value
        if status_value == "Đã chấm":
            status_cell.font = Font(color="008000", bold=True)
        elif status_value == "Cần xem lại":
            status_cell.font = Font(color="FF8C00", bold=True)
        elif status_value == "Đang xử lý":
            status_cell.font = Font(color="0000FF")
        elif status_value == "Chấm tay":
            status_cell.font = Font(color="800080", bold=True)

    # ===== ĐỊNH DẠNG CỘT SỐ CÂU ĐÚNG =====
    for row in range(2, len(results) + 2):
        correct_cell = ws.cell(row=row, column=5)
        if correct_cell.value is not None:
            try:
                correct = int(correct_cell.value)
                total_cell = ws.cell(row=row, column=6)
                total = int(total_cell.value) if total_cell.value else max_q
                if total > 0:
                    ratio = correct / total
                    if ratio >= 0.8:
                        correct_cell.font = Font(color="008000", bold=True)
                    elif ratio >= 0.5:
                        correct_cell.font = Font(color="0000FF")
                    else:
                        correct_cell.font = Font(color="FF0000")
            except:
                pass

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    ws.freeze_panes = 'C2'

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    filename = f"BangDiem_{session_obj.session_code}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
